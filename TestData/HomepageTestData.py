import openpyxl

class TestDataHomePage:

    test_case2_Data = [{"Name":"Luffy","Email":"luffy@gmail.com","Password":"23456"},
                       {"Name":"Zoro","Email": "zoro@gmail.com","Password":"56789"}]

    @staticmethod
    def getData(testcase_name):

        book = openpyxl.load_workbook(r'C:\\Users\eshwar.m\Downloads\PythonData.xlsx')
        Dict = {}
        sheet = book.active

        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == testcase_name:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [Dict]